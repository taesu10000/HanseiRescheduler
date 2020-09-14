from django.shortcuts import render, redirect
from .models import ScheduleInfo, ClassesInfo, Document
from .forms import UploadFileForm
from django.http import HttpResponse
from datetime import *
from django.contrib.auth.models import User
from django.contrib import auth
from django.contrib.auth.hashers import check_password
import openpyxl, os, json
from openpyxl import load_workbook



def signinpage(request):
    print("signinpage")
    if (request.user.is_authenticated is False): #로그인 안되어 있으면
        return render(request, 'HanseiApp/signin.html')
    else :
        return render(request, 'HanseiApp/index.html')
        
def loginpage(request):
    print("loginpage")
    if (request.user.is_authenticated is False or request.user.is_anonymous is True): #로그인 안되어 있으면, 로그인없이 사용하는자라면
        return render(request, 'HanseiApp/login.html')
    else :
        return render(request, 'HanseiApp/index.html')
        
def params(request):
    print("params")
    if (request.user.is_authenticated is False or request.user.is_anonymous is True): #로그인 안되어 있으면, 로그인없이 사용하는자라면
        return render(request, 'HanseiApp/login.html')
    else :
        return render(request, 'HanseiApp/params.html')


def examples(request):
    print("examples")
    return render(request, 'HanseiApp/examples.html')

def signin(request):
    # DB Insert
    print("signin")
    if request.method == "POST":
        context = {}
        if request.POST['inputPassword'] == request.POST['inputPassword2']:
            try:
                user = User.objects.create_user(username=request.POST['inputEmail'], email=request.POST['inputEmail'], password=request.POST['inputPassword'])
                auth.login(request, user)
                return render(request, 'HanseiApp/index.html')
            except:
                if IsUserExist(request.POST['inputEmail']):
                    context = {'message':"The E-mail you want to use is already used by anther user."}
                else:
                    context = {'message':"Please Check your E-mail. If you keep receiving this message, Ask Kim."}
            return render(request, 'HanseiApp/signin.html', context) 
        context = {'message':"Please Check your Password. which doens't match."}
        return render(request, 'HanseiApp/signin.html', context) 
    return HttpResponse("You entered through wrong way! Ask Kim")

def login(request):
    # Get from DB
    print("login")
    if request.method == "POST":
        Email_Address = request.POST['inputEmail'] # username
        Pass = request.POST['inputPassword']
        user = auth.authenticate(request, username=Email_Address, password=Pass)
        IsusernameExists = IsUserExist(Email_Address)
        context = {}
        if  user is not None:
            auth.login(request, user)
            #쿠키 설정
            context = {'user':user}
            response = render(request, 'HanseiApp/index.html',context)
            response.set_cookie('username',Email_Address)
            response.set_cookie('password',Pass)
            return response
        else:
            if (IsusernameExists is False):
                context = {'message':"Please Check your E-mail"}
            else:
                context = {'message':"Please Check your Password"}
        return render(request, 'HanseiApp/login.html', context) 
    return HttpResponse("You entered through wrong way! Ask Kim")

def logout(request):
    #쿠키 제거
    print("logout")
    response = render(request, 'HanseiApp/login.html')
    response.delete_cookie('username')
    response.delete_cookie('password')
    auth.logout(request)
    return response

def index(request):
    print("index")
    if (request.user.is_authenticated is False or request.user.is_anonymous is True): #로그인 안되어 있으면, 로그인없이 사용하는자라면
        return render(request, 'HanseiApp/login.html')
    else :
        user = request.user
        context = {'user':user}
        return render(request, 'HanseiApp/index.html', context) 
 

def manual(request):
    print("manual")
    if (request.user.is_authenticated is False or request.user.is_anonymous is True):#로그인이 되어 있으면,
        return render(request, 'HanseiApp/login.html')


    return render(request, 'HanseiApp/manual.html')

def timetable(request):
    print("timetable")
    if (request.user.is_authenticated is False or request.user.is_anonymous is True):#로그인이 되어 있으면,
        return render(request, 'HanseiApp/login.html')
        
    #DB 읽어서 표현하는 부분
    context = GetContext(request)  #오늘 날짜와 본인의 정보를 받아옴
    print("Recieved")
    print(request.POST)
    print("context")
    print(context)
    return render(request, 'HanseiApp/timetable.html', context)
        #return HttpResponse("ERROR") 
    #DB 읽어서 표현하는 부분

def timetableSubmit(request): 
    print("timetableSubmit")
    if (request.user.is_authenticated is False or request.user.is_anonymous is True):#로그인이 되어 있으면,
        return render(request, 'HanseiApp/login.html')

    if request.method == "POST" and 'schedule' in request.POST: 
        ##################Post 값 저장 등록 & 업데이트################## 

        TimetableCellInfo = request.POST['timetablecellinfo']
        Memo = request.POST['memo']
        is_staff = request.user.is_staff

        priky = GetPk(request) #PrimaryKey = 년도+1년중주수(52주)+유저PK
        infos = ScheduleInfo.objects.all()
        infoExists = False
        for info in infos:         #info가 있을 경우 수정
            if  info.pk == priky:
                info.Inputdate = datetime.today().strftime("%Y-%m-%d") #DB에는 현재 등록 시간을 기록
                info.TimetableCellInfo = TimetableCellInfo
                info.user = request.user
                info.Memo = Memo
                infoExists = True
                info.save()
                
        if infoExists is False:     #info가 없을 경우
            info = ScheduleInfo(pk=priky, TimetableCellInfo=TimetableCellInfo, Inputdate=datetime.today().strftime("%Y-%m-%d"), User=request.user, Memo=Memo)
            info.save()

    print("Recieved")
    print(request.POST)
    #print("context")
    context = GetContext(request)
    #print(context)

    return render(request, 'HanseiApp/timetable.html', context)
        #return HttpResponse("ERROR") 
    #DB 읽어서 표현하는 부분

def classinfo(request):
    print("classinfo")
    if (request.user.is_authenticated is False or request.user.is_anonymous is True):#로그인이 되어 있으면,
        return render(request, 'HanseiApp/login.html')

 
    if request.method == "POST":
        print("request.POST")
        print(request.POST)
    #     infos = ClassesInfo.objects.all()
    #     for info in infos:   
    #         info.delete()
    #     try:
    #         info = ClassesInfo(ClassInfo=request.POST['classes'])
    #         info.save()
    #     except:
    #         pass
    # context = {}
    # infos = ClassesInfo.objects.all()
    # for info in infos:
    #     context = {'ClassInfo':info.ClassInfo}
    context = {}
    print("context")
    infos = Document.objects.all()
    for info in infos:
        context['ImpertedExcel'] = info.ImpertedExcel
        #print(context['ClassInfo'])
    infos = ClassesInfo.objects.all()
    for info in infos:
        context['ClassInfo'] = info.ClassInfo
        print(context['ClassInfo'])
    
    #print(context['ImpertedExcel'])
    return render(request, 'HanseiApp/classinfo.html', context)

def GetContext(request):
    #request.user > 현재 접속자
    curDatestr = GetcurSelDay(request)   #select된 날짜있으면, 아니면 오늘 날짜 반환
    priky = GetPk(request) #PrimaryKey = 년도+1년중주수(52주)+유저PK
    AllUsers = GetAllUsers()
    seluser = GetUserFormat(request) #UserName + Userpk + Isstaff
    TimetableCellInfo = ""
    Scheduleinfos = ScheduleInfo.objects.all()
    Memo = ""
    for info in Scheduleinfos:
        if info.pk == priky: #보여 줘야 할 정보인지
            TimetableCellInfo = info.TimetableCellInfo
            Memo = info.Memo
    Classes = ""
    ImpertedExcel = ""
    Classesinfos = ClassesInfo.objects.all()
    for info in Classesinfos:
        Classes = info.ClassInfo
    infos = Document.objects.all()
    for info in infos:
        ImpertedExcel = info.ImpertedExcel
    context = {'curDatestr':curDatestr, 'TimetableCellInfo':TimetableCellInfo, 'seluser':seluser, 'AllUsers':AllUsers,'ImpertedExcel':ImpertedExcel, 'ClassesInfo':Classes, 'Memo':Memo} 
    
    return context

def GetPk(request):  #PrimaryKey = 년도+1년중주수(52주)+유저PK is_staff가 True일때는 웹에서 선택 된 사람의 PK를 가져 와야 한다. 아닐때는 접속자의PK
    curSelDay = GetcurSelDay(request)
    pk = request.user.pk
    if (request.user.is_staff == True):
        if request.method == "POST":
            try:
                pk = request.POST['User_selection']
                if (pk is None):
                    print(pk)
                    pk = request.user.pk
            except:
                pk = request.user.pk
    return int(format(str(curSelDay[0:4]) + str(GetMonthNo(curSelDay)) + str(pk)))

def GetMonthNo(currentweek):
    curweek = datetime(int(currentweek[0:4]), int(currentweek[5:7]), int(currentweek[8:10])) 
    month = curweek.isocalendar()
    return month[1]

def GetcurSelDay(request): #Request가 있으면 select된 날짜 아니면 오늘 날짜 반환
    if 'Search' in request.POST or 'SelectedDate' in request.POST: 
        return request.POST['SelectedDate']
    else:
        return datetime.today().strftime("%Y-%m-%d")

def GetUserFormat(request):
    is_staff = request.user.is_staff
    seluser = request.user.username  #사용자
    userPk = request.user.pk
    if (is_staff == True): # 매니져 로그인 했을때
        try:
            seluser = request.POST['seluser']# 선택된 사람
            if(seluser is None):
                seluser = request.user.username
        except:
            pass

    return seluser+"$"+str(userPk)+"$"+str(is_staff)

def IsUserExist(InputUserName):
    FoundUser = False
    infos = User.objects.all()
    for info in infos:  
        if(info.username == InputUserName):
            FoundUser = True
            break
    return FoundUser

def GetAllUsers():
    userlist = ""
    infos = User.objects.all()
    for info in infos:
        if (info.is_superuser == False):
            if userlist == "": 
                userlist += info.username+"$"+str(info.pk)
            else:
                userlist += "|"+ info.username+"$"+str(info.pk)
    return userlist

def timetableimport(request):
    print("timetableimport")
    if request.method == "POST":
        print("request.FILES")
        print(request.FILES)
        print("request.POST")
        print(request.POST)
        excel_file = ""
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = handle_uploaded_file(request.FILES['Excelimport'],request.POST['FileName']) #엑셀파일이 열려 있을 경우 열리지 않음 에러출력
        else:
            print("is_not_valid")
            form = UploadFileForm()
            return render(request, 'HanseiApp/timetableimport.html')
        Timetable = {}
        Classes = {}
        GetExcelTableInfo(request, excel_file, Timetable, Classes)
        print(Timetable)
        infos = Document.objects.all()
        for info in infos:
            info.delete()
        try:
            info = Document(ImpertedExcel=json.dumps(Timetable))
            info.save()
        except:
            pass

        infos = ClassesInfo.objects.all()
        for info in infos:
            info.delete()
        try:
            info = ClassesInfo(ClassInfo=json.dumps(Classes))
            info.save()
        except:
            pass
        context = {'message' : 'Uploaded!'}
        return render(request, 'HanseiApp/timetableimport.html', context)
    return render(request, 'HanseiApp/timetableimport.html')

def handle_uploaded_file(excel_file, fileName):
    path = "C:\\Hansei\\Downloaded\\"
    try:
        if not(os.path.isdir(path)):
            os.makedirs(os.path.join(path))
    except OSError as e:
        if e.errno != errno.EEXIST:
            print("Failed to create directory!!!!!")
    try:
        with open(path+fileName, 'wb+') as destination:
            for chunk in excel_file.chunks():
                destination.write(chunk)
    except :
        print("Failed to create directory!!!!!")

    try:
        if not(os.path.isdir(path)):
            os.rmdir(os.path.join(path))
    except OSError as e:
        if e.errno != errno.EEXIST:
            print("Failed to delete directory!!!!!")

    return path+fileName

def GetExcelTableInfo(request, excel_file, Timetable, Classes): #영역의 모든 셀을 다 돌기 때문에 부하가 걸릴 위험이 큰 구간
    wb = load_workbook(excel_file, data_only = True)
    shName = wb.get_sheet_names()
    sheetName = request.POST['sheetName']
    MaxRow = request.POST['sheetRowCnt']
    MaxCol = request.POST['sheetCloCnt']
    if sheetName in shName:
        sh = wb[sheetName]

        ClassesNameKey = 0        
        ClasseRoom = {}
        for r in range(1, int(MaxRow)):
            for c in range(1, int(MaxCol)):
                cell = sh.cell(row=r, column= c)

                if(cell.value == "Key for classes"):    #"Key for classes" 찾기
                    key = "key"
                    ClassesNameKey = r
                if(cell.fill.start_color.index == "FFFFFF00" and cell.value is not None): #강의실 이름 찾기  FFFFFF00 > 노란색
                    key = '%d:%d' %(r,c)
                    ClasseRoom[key] = cell.value

        GetClassLevel(sh, ClassesNameKey, Classes)  #"Key for classes"하위 클래스들을 이름:색상 정보를 포함하여 가져옴
        GetClassTable(sh, ClasseRoom, Timetable)          #모든 "노란색" 셀(강의실)의 시간표를 가져옴
   
def GetClassTable(sh, ClassRooms, Timetable):
    if(ClassRooms is not None):
        #Timetable = {}
        for pos in ClassRooms:
            #print("pos : %s ClassRooms : %s" %(pos, ClassRooms[pos]))
            postime = pos[0:pos.find(":")]
            posday = pos[pos.find(":")+1:len(pos)]
            class_ = {}
            #print("postime : %s posday : %s" %(postime, posday))
            for time in range(1,11): #1~10교시까지
                for day in range(1,6): # 5일 월화수목금 정보에 접근
                    class_prop_ = {}
                    row = time+int(postime)
                    col = day+int(posday)
                    if (row > 0 and col > 0):
                        cell = sh.cell(row=row, column=col)
                        value = cell.value
                        color = cell.fill.start_color.index
                        if (value != None or (color != "FFFFFFFF" and color != "00000000")):#강의실(시간표 좌상단)기준으로 표의 정보를 가져옴 색상기준 애매모호
                            propkey = "%s%d" %(Getday(day), time+8)
                            class_prop_[value] = "#"+color[2:8]
                            class_[propkey]= class_prop_
            Timetable[ClassRooms[pos]] = class_

def Getday(day):
    strDay = ""
    if(day == 1):
        strDay = "mon"
    elif(day == 2):
        strDay = "tue"
    elif(day == 3):
        strDay = "wed"
    elif(day == 4):
        strDay = "thu"
    elif(day == 5):
        strDay = "fri"
    return strDay

def GetClassLevel(sheet, foundKeyPos, Classes):
    if (foundKeyPos != 0):
        for r in range(foundKeyPos+1,sheet.max_row):        
            if(sheet.cell(row=r,column=1).value != None):
                color = sheet.cell(row=r,column=1).fill.start_color.index
                value = sheet.cell(row=r,column=1).value
                key = "%d:1" % r 
                Classes[key] = "%s|#%s" %(value, color[2:8]) #위치, 값, 컬러 
            else:
                break
    # for key in ClassKey:
    #     print(key +" "+ ClassKey[key])     


#규칙 1. 반드시 A열에 있어야 검색이 가능함
#규칙 2. 강의실 이름은 반드시 FFFFFF00 색으로 통일 되어야 함
#규칙 3. 빈 강의실 색 FF666666
#규칙 4. 강의실 이름을 기준으로 5칸(요일)컬럼과 10개(교시)의 테이블 범위를 만들어 낸다.


# def GetClassTable(sh, ClassRooms, Timetable):
#     if(ClassRooms is not None):
#         #Timetable = {}
#         for pos in ClassRooms:
#             #print("pos : %s ClassRooms : %s" %(pos, ClassRooms[pos]))
#             postime = pos[0:pos.find(":")]
#             posday = pos[pos.find(":")+1:len(pos)]
#             class_ = {}
#             #print("postime : %s posday : %s" %(postime, posday))
#             for time in range(1,11): #1~10교시까지
#                 for day in range(1,6): # 5일 월화수목금 정보에 접근
#                     class_prop_ = {}
#                     row = time+int(postime)
#                     col = day+int(posday)
#                     if (row > 0 and col > 0):
#                         cell = sh.cell(row=row,column=col)
#                         value = cell.value
#                         color = cell.fill.start_color.index
                        
#                         if (value != None or (color != "FFFFFFFF" and color != "00000000")):#강의실(시간표 좌상단)기준으로 표의 정보를 가져옴 색상기준 애매모호
#                             propkey = "%s%d" %(Getday(day),time+8)
#                             class_prop_[value] = color
#                             class_[propkey]= class_prop_                        
#             Timetable[ClassRooms[pos]] = class_



        # for col in sheet.columns:
        #     if "Key for classes" in sheet:
        #         #print("Key for classes")
        #         #print(sheet.rows+ " " +sheet.cols)
        #         foundKey = True
        #         break
        # if foundKey == True:
        #   break

# def UpdateDB():   
#     try :
#     infos = ScheduleInfo.objects.all()
#     for info in infos:         #info가 있을 경우 수정
#         if  info.pk == priky:
#             info.Inputdate = datetime.today().strftime("%Y-%m-%d") #DB에는 현재 등록 시간을 기록
#             info.AvailableTime = AvailableTime
#             info.user = request.user
#             infoExists = True
#             info.save()
            
# def parse(string):
#     return str(string).replace('\'', '').replace('[', '').replace(']', '')

# def GetDayNo(year, month, day): #날짜 입력 하면 요일 반환 함수
#     currentweek = date(year, month, day).weekday()
#     return currentweek
#     #0 Mon #3 Tur #7 Sun 
#     #1 Tue #4 Fri
#     #2 Wed #5 Sat

# def GetWeekNo(currentDate): #이번달 중 몇번째 주인지 구하는 함수
#     firstday = currentDate.replace(day=1)
#     if firstday.weekday() == 6:
#         origin = firstday
#     elif firstday.weekday() < 3:
#         origin = firstday - timedelta(days=firstday.weekday() + 1)
#     else:
#         origin = firstday + timedelta(days=6-firstday.weekday())
#     return (currentDate - origin).days // 7 + 1


#def blabla():
    # for key in ClassesKey:
    #     print(ClassesKey +" "+ ClassesKey[key])

    # for key in ClassKey
    #     print(key)
    #GetClassKey(sh)
    #color_in_hex = sh['A9'].fill.start_color.index
    #print(color_in_hex)
    # shValue = sh['A9'].value
    # print ("HEX : " + color_in_hex + " shValue : " + shValue) 
    # print(sh.max_row)
    #color_in_hex = sh['A3'].fill.start_color.index # this gives you Hexadecimal value of the color
    #print ('HEX =',color_in_hex) 
    #print('RGB =', tuple(int(color_in_hex[i:i+2], 16) for i in (0, 2, 4))) # Color in RGB
